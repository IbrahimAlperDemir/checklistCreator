from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from form_tracker import get_next_form_number

def save_to_checklist(text: str, filename: str, revision: str = "A") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Checklist"

    # 🔢 Form numarası
    form_number = get_next_form_number("testcase")

    # 📋 Üst Bilgi Satırları
    ws.append(["Form Adı", "Test Senaryo Kontrol Listesi"])
    ws.append(["Form Numarası", f"FRM-TST-{form_number}"])
    ws.append(["Revizyon", revision])
    ws.append(["Tarih", datetime.today().strftime("%d.%m.%Y")])
    ws.append([])  # boş satır

    # 📌 Başlık Satırı (güncellenmiş sütun isimleri)
    ws.append(["Test Case No", "Test Adımı", "Beklenen Sonuç", "Ön Koşul", "Test Tipi"])

    # 📄 İçerik: "Test Adımları ve Beklenen Sonuçlar" kısmından itibaren işlenir
    processing = False
    for line in text.splitlines():
        if "Test Adımları ve Beklenen Sonuçlar" in line:
            processing = True
            continue
        if processing and line.strip():
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 5:
                ws.append(parts[:5])
            else:
                ws.append(parts + [""] * (5 - len(parts)))

    # 🎨 Stil: Üst bilgi + başlık satırları kalın ve ortalanmış
    for row in ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # 🔧 Sütun genişlikleri otomatik ayarla
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_length + 5

    # 💾 Kaydet
    wb.save(filename)
    print(f"✅ '{filename}' başarıyla kaydedildi.")
